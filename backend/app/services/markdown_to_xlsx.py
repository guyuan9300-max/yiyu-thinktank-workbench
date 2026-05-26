"""Markdown → .xlsx 渲染器 (openpyxl)。

工作台 chat 答案导出策略:
  · 主要是表格 + 标题 (无大段文字) → 这个 renderer, 真返 xlsx
  · 有大段说明文字 → markdown_to_docx 那个

xlsx 真布局:
  · 每个 markdown 表格 → 一个独立 sheet
  · 第 1 行 = headers (加粗 + 浅蓝填充, header 视觉锁定)
  · 多个表格时: sheet 名 "表格1" / "表格2" / ... (LLM 真给的 H2/H3 标题不可靠, 用序号稳)
  · 如果非表格部分有标题 (#) 或少量文字, 真额外加一个 "说明" sheet 真写它们
  · cell 内 markdown 内联装饰 (**加粗** / *斜体* / [链接](url)) 真去掉, 转纯文本
    (Excel cell 不支持多 run / 富文本, 真接受降级)
"""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from app.services.markdown_table import (
    MarkdownTable,
    markdown_inline_to_plain,
    parse_markdown_tables,
    strip_tables,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


_HEADER_FILL_COLOR = "DCE4FC"  # 真浅蓝, 跟主站 #5B7BFE 同色系但更浅
_HEADER_FONT_COLOR = "1E3A8A"  # 真深蓝


def _autosize_columns(ws: "Worksheet", *, max_width: int = 100) -> None:
    """真按内容长度自动调列宽 (粗略, 不超过 max_width)。
    顾源源 5/26: 60 → 100 真宽点, '详情' 列里长文本不挤。"""
    for column_cells in ws.columns:
        max_len = 0
        column_letter = None
        for cell in column_cells:
            if column_letter is None:
                column_letter = cell.column_letter
            value = "" if cell.value is None else str(cell.value)
            # 真中文字符按 1.7 倍宽 (Excel 默认西文字符宽)
            visual_width = sum(1.7 if ord(ch) > 127 else 1.0 for ch in value)
            if visual_width > max_len:
                max_len = visual_width
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_len + 2, max_width)


def _write_table_to_sheet(ws: "Worksheet", table: MarkdownTable) -> None:
    """真把一个 markdown 表格写到 sheet。"""
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color=_HEADER_FONT_COLOR, name="PingFang SC")
    header_fill = PatternFill(start_color=_HEADER_FILL_COLOR, end_color=_HEADER_FILL_COLOR, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="PingFang SC")
    body_align = Alignment(vertical="center", wrap_text=True)

    # 真写表头
    for col_idx, header in enumerate(table.headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=markdown_inline_to_plain(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    # 真写 body
    for row_idx, row in enumerate(table.rows, start=2):
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=markdown_inline_to_plain(cell_value))
            cell.font = body_font
            cell.alignment = body_align
    # 真冻结首行 (header 视觉锁定)
    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _write_notes_to_sheet(ws: "Worksheet", markdown_text: str) -> None:
    """真把非表格部分写到 '说明' sheet。
    顾源源 5/26: 真跳过 H 标题行 (它们已经做 sheet name, 重复写真冗余),
    只写真"正文段落 + 非 heading 真内容"。
    """
    from openpyxl.styles import Alignment, Font

    body_font = Font(name="PingFang SC")
    align = Alignment(vertical="top", wrap_text=True)

    row_idx = 1
    for raw_line in markdown_text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        # 真跳过 heading 行 (已做 sheet name, 不重复写)
        if line.startswith("#"):
            continue
        # 真跳过纯分隔线
        import re as _re
        if _re.match(r"^[-=]{3,}$", line):
            continue
        # 真普通段落 (去 markdown 内联 + evidence badge)
        cell = ws.cell(row=row_idx, column=1, value=markdown_inline_to_plain(line))
        cell.font = body_font
        cell.alignment = align
        row_idx += 1

    ws.column_dimensions["A"].width = 100


def _sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    """Excel sheet 名真规则:
      · ≤ 31 字符
      · 真禁用字符: \\ / ? * [ ] :
      · 不能跟已有 sheet 名重名 → 加 ' (2)' 后缀
    """
    # 真去禁用字符
    cleaned = name
    for ch in ["\\", "/", "?", "*", "[", "]", ":"]:
        cleaned = cleaned.replace(ch, "")
    cleaned = cleaned.strip() or "Sheet"
    # truncate 到 31 字符
    base = cleaned[:31]
    final = base
    counter = 2
    while final in used_names:
        suffix = f" ({counter})"
        final = (base[: 31 - len(suffix)]) + suffix
        counter += 1
    return final


def render_markdown_to_xlsx_bytes(
    markdown_text: str,
    *,
    document_title: str | None = None,
) -> bytes:
    """把 markdown 真渲染成 .xlsx 字节流。

    布局 (顾源源 5/26 真升级):
      · 每个 markdown 表格 → 一个独立 sheet
        sheet 名优先用真表格上方最近的 # 标题 (去序号前缀, 例: "一、机构基础信息表" → "机构基础信息表")
        没找到 fallback 用"表格N"
      · 非表格部分仅在真有"非 heading 非空白" 内容时才建"说明" sheet
        (避免"说明"里只剩跟 sheet 名重复的 H 标题)
      · 真无表格时, 全部内容写到 "说明" sheet
    """
    # 真延迟 import (避免 backend 启动被 openpyxl 拖)
    from openpyxl import Workbook

    wb = Workbook()
    # 真删 openpyxl 默认建的 "Sheet"
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    tables = parse_markdown_tables(markdown_text)
    non_table_text = strip_tables(markdown_text)
    used_sheet_names: set[str] = set()

    # 真先建表格 sheets (用真 heading 当 sheet 名)
    if tables:
        for idx, table in enumerate(tables, start=1):
            raw_name = table.heading_above or (f"表格{idx}" if len(tables) > 1 else "表格")
            sheet_name = _sanitize_sheet_name(raw_name, used_sheet_names)
            used_sheet_names.add(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            _write_table_to_sheet(ws, table)

    # 真判非表格部分是否值得建 "说明" sheet
    # 顾源源 5/26: 真排除"只剩 # 标题 + 空行" 真无意义内容 (那些 heading 已经做 sheet 名了)
    # 真真有内容 = 至少 1 行真非空 + 非 heading 的真正文本
    import re as _re

    has_meaningful_notes = False
    for line in non_table_text.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        # 真跳过纯 heading 行 (# / ## / ###)
        if _re.match(r"^#{1,5}\s+", stripped):
            continue
        # 真跳过纯分隔线
        if _re.match(r"^[-=]{3,}$", stripped):
            continue
        has_meaningful_notes = True
        break

    if has_meaningful_notes:
        notes_ws = wb.create_sheet(title=_sanitize_sheet_name("说明", used_sheet_names))
        _write_notes_to_sheet(notes_ws, non_table_text)

    # 真兜底: 真完全空, 至少建一个空 sheet
    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="空")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
